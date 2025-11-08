import sys
import re
import os
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


VENUES = ["札幌","函館","福島","新潟","東京","中山","中京","京都","阪神","小倉"]

# ===== ユーティリティ =====
def anchor_text(cell: Tag) -> str:
    if isinstance(cell, Tag):
        a = cell.find("a")
        if a:
            t = a.get_text(strip=True)
            if t:
                return t
        return cell.get_text(strip=True)
    return str(cell).strip()

def clean_name(s: str) -> str:
    s = re.split(r"[（(]", s, maxsplit=1)[0]
    return s.strip()

def find_col_index(header_map: dict, candidates) -> int | None:
    # 完全一致→部分一致
    for key in candidates:
        for h, idx in header_map.items():
            if h == key:
                return idx
    for key in candidates:
        for h, idx in header_map.items():
            if key in h:
                return idx
    return None

# ===== 出馬表テーブル検出 =====
def find_table_and_headers(soup: BeautifulSoup):
    for t in soup.find_all("table"):
        thead = t.find("thead")
        if thead:
            head_cells = thead.find_all(["th", "td"])
        else:
            first_tr = t.find("tr")
            head_cells = first_tr.find_all(["th","td"]) if first_tr else []
        if not head_cells:
            continue

        heads_raw = [hc.get_text(strip=True) for hc in head_cells]
        heads_norm = [re.sub(r"\s+","", h) for h in heads_raw]
        has_horse = any("馬名" in h for h in heads_norm)
        has_jock  = any(("騎手" in h) or ("騎手名" in h) for h in heads_norm)
        if has_horse and has_jock:
            header_map = {h:i for i,h in enumerate(heads_norm)}
            return t, heads_raw, header_map
    return None, None, None

# ===== 帯（ページ上部）の基本メタ：日付・場所・R =====
def extract_basic_meta(text_all: str):
    # 日付
    m_date = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text_all)
    if m_date:
        y, mo, d = map(int, m_date.groups())
        ymd = f"{y:04d}{mo:02d}{d:02d}"
    else:
        ymd = datetime.now().strftime("%Y%m%d")

    # 「○回[場所]○日」から場所を厳密抽出
    m_place = re.search(r"\d+\s*回\s*(%s)\s*\d+\s*日" % "|".join(VENUES), text_all)
    place = m_place.group(1) if m_place else "不明"

    # レース番号（8レース or 8R）
    m_r1 = re.search(r"(\d{1,2})\s*レース", text_all)
    m_r2 = re.search(r"(\d{1,2})\s*R", text_all)
    if m_r1:
        race_no = f"{int(m_r1.group(1))}R"
    elif m_r2:
        race_no = f"{int(m_r2.group(1))}R"
    else:
        race_no = "R"
    return ymd, place, race_no

# ===== レース名：class="race_name" を最優先 =====
def extract_race_title(soup: BeautifulSoup, fallback_place: str, fallback_race_no: str) -> str:
    # 1) まず .race_name をピンポイントで取得
    rn = soup.select_one(".race_name")
    if rn:
        t = rn.get_text(strip=True)
        if t:
            # パイプ等の右側ノイズがあれば左側だけ
            return t.split("|")[0].strip()

    # 2) フォールバック：<title> から左側を取り、余計な語が混ざれば最終フォールバックへ
    title_tag = soup.find("title")
    if title_tag:
        t = title_tag.get_text(strip=True).split("|")[0].strip()
        if t:
            return t

    # 3) 最終フォールバック
    return f"{fallback_place}{fallback_race_no}"

# ===== 本体取得 =====
def fetch_rows_and_meta(url: str):
    r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=15)
    r.raise_for_status()
    r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text, "lxml")

    table, heads, header_map = find_table_and_headers(soup)
    if not table:
        raise SystemExit("出馬表テーブル（ヘッダーに『馬名』『騎手』）を検出できませんでした。")

    col_umaban = find_col_index(header_map, ["馬番","馬番号"])
    col_horse  = find_col_index(header_map, ["馬名"])
    col_jock   = find_col_index(header_map, ["騎手","騎手名"])
    if col_horse is None or col_jock is None:
        raise SystemExit("『馬名』『騎手(騎手名)』の列を特定できませんでした。")

    trs = table.find_all("tr")
    start_idx = 1 if trs and trs[0].find_all("th") else 0

    rows = []
    for tr in trs[start_idx:]:
        tds = tr.find_all(["td","th"])
        if len(tds) <= max(col_horse, col_jock):
            continue

        horse = clean_name(anchor_text(tds[col_horse]))
        jockey = clean_name(anchor_text(tds[col_jock]))
        if not horse or re.fullmatch(r"\d+", horse):  # 馬名が数字のみは除外
            continue
        if not jockey or jockey == "-":
            continue

        # 馬番
        umaban = ""
        if col_umaban is not None and len(tds) > col_umaban:
            um_raw = anchor_text(tds[col_umaban]).strip()
            m = re.search(r"\d{1,2}", um_raw)
            if m:
                umaban = m.group(0)
        else:
            maybe = anchor_text(tds[0]) if tds else ""
            m = re.match(r"\D*(\d{1,2})\D*", maybe)
            if m:
                umaban = m.group(1)

        rows.append((umaban, horse, jockey))

    if not rows:
        raise SystemExit("馬番／馬名／騎手名の抽出結果が空でした。")

    # ファイル名用メタ
    text_all = soup.get_text(" ", strip=True)
    ymd, place, race_no = extract_basic_meta(text_all)

    # レース名は .race_name を最優先で取得
    race_title = extract_race_title(soup, place, race_no)

    filename = f"{ymd}_{place}_{race_no}.xlsx"
    return rows, filename, race_title

# ===== Excel 出力 =====
def save_to_desktop(rows, filename, race_title):
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    os.makedirs(desktop, exist_ok=True)
    path = os.path.join(desktop, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "出馬表"

    # 見出し（レース名）
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = race_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True, size=18)
    title_cell.fill = PatternFill(start_color="FADADD", end_color="FADADD", fill_type="solid")
    ws.row_dimensions[1].height = 30  # ≈ 30px

    # ヘッダー
    headers = ["馬番","馬名","騎手名","評価","短評"]
    ws.append(headers)

    # データ
    for umaban, horse, jockey in rows:
        ws.append([umaban, horse, jockey, "", ""])

    # 体裁
    light_blue = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    bold = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    max_row = ws.max_row
    max_col = 5

    # ヘッダー装飾
    for c in range(1, max_col + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = light_blue
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 表全体：中央寄せ＆罫線
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列幅
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50

    wb.save(path)
    return path

# ===== main =====
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python jra_card_cli.py <出馬表URL>")
        sys.exit(1)

    url = sys.argv[1].strip()
    rows, filename, race_title = fetch_rows_and_meta(url)
    out_path = save_to_desktop(rows, filename, race_title)
    print(f"✅ 保存完了: {out_path}")