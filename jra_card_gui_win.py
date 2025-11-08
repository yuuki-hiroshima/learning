import os
import re
import threading
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox

import requests
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


VENUES = ["札幌","函館","福島","新潟","東京","中山","中京","京都","阪神","小倉"]


# ====== 共通ユーティリティ ======
def anchor_text(cell: Tag) -> str:
    if isinstance(cell, Tag):
        a = cell.find("a")
        if a:
            t = a.get_text(strip=True)
            if t:
                return t
        return cell.get_text(strip=True)
    return str(cell).strip()

def clean_name(s: str) -> str:
    s = re.split(r"[（(]", s, maxsplit=1)[0]
    return s.strip()

def find_col_index(header_map: dict, candidates) -> int | None:
    for key in candidates:
        for h, idx in header_map.items():
            if h == key: return idx
    for key in candidates:
        for h, idx in header_map.items():
            if key in h: return idx
    return None

def find_table_and_headers(soup: BeautifulSoup):
    for t in soup.find_all("table"):
        thead = t.find("thead")
        if thead:
            head_cells = thead.find_all(["th", "td"])
        else:
            first_tr = t.find("tr")
            head_cells = first_tr.find_all(["th","td"]) if first_tr else []
        if not head_cells:
            continue

        heads_raw = [hc.get_text(strip=True) for hc in head_cells]
        heads_norm = [re.sub(r"\s+","", h) for h in heads_raw]
        has_horse = any("馬名" in h for h in heads_norm)
        has_jock  = any(("騎手" in h) or ("騎手名" in h) for h in heads_norm)
        if has_horse and has_jock:
            header_map = {h:i for i,h in enumerate(heads_norm)}
            return t, heads_raw, header_map
    return None, None, None

def extract_basic_meta(text_all: str):
    m_date = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text_all)
    if m_date:
        y, mo, d = map(int, m_date.groups())
        ymd = f"{y:04d}{mo:02d}{d:02d}"
    else:
        ymd = datetime.now().strftime("%Y%m%d")

    # 例：「4回京都1日」から厳密取得
    m_place = re.search(r"\d+\s*回\s*(札幌|函館|福島|新潟|東京|中山|中京|京都|阪神|小倉)\s*\d+\s*日", text_all)
    place = m_place.group(1) if m_place else "不明"

    m_r1 = re.search(r"(\d{1,2})\s*レース", text_all)
    m_r2 = re.search(r"(\d{1,2})\s*R", text_all)
    if m_r1:
        race_no = f"{int(m_r1.group(1))}R"
    elif m_r2:
        race_no = f"{int(m_r2.group(1))}R"
    else:
        race_no = "R"
    return ymd, place, race_no

def fetch_rows_and_meta_by_url(target_url: str):
    # 取得
    r = requests.get(target_url, headers={"User-Agent":"Mozilla/5.0"}, timeout=15)
    r.raise_for_status()
    r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text, "lxml")

    # 出馬表テーブル
    table, heads, header_map = find_table_and_headers(soup)
    if not table:
        raise RuntimeError("出馬表テーブル（『馬名』『騎手』ヘッダー）を検出できませんでした。URLをご確認ください。")

    col_umaban = find_col_index(header_map, ["馬番","馬番号"])
    col_horse  = find_col_index(header_map, ["馬名"])
    col_jock   = find_col_index(header_map, ["騎手","騎手名"])
    if col_horse is None or col_jock is None:
        raise RuntimeError("『馬名』『騎手(騎手名)』列を特定できませんでした。")

    trs = table.find_all("tr")
    start_idx = 1 if trs and trs[0].find_all("th") else 0

    rows = []
    for tr in trs[start_idx:]:
        tds = tr.find_all(["td","th"])
        if len(tds) <= max(col_horse, col_jock):
            continue

        horse = clean_name(anchor_text(tds[col_horse]))
        jockey = clean_name(anchor_text(tds[col_jock]))
        if not horse or re.fullmatch(r"\d+", horse):
            continue
        if not jockey or jockey == "-":
            continue

        umaban = ""
        if col_umaban is not None and len(tds) > col_umaban:
            um_raw = anchor_text(tds[col_umaban]).strip()
            m = re.search(r"\d{1,2}", um_raw)
            if m:
                umaban = m.group(0)
        else:
            maybe = anchor_text(tds[0]) if tds else ""
            m = re.match(r"\D*(\d{1,2})\D*", maybe)
            if m:
                umaban = m.group(1)

        rows.append((umaban, horse, jockey))

    if not rows:
        raise RuntimeError("馬番／馬名／騎手名の抽出結果が空でした。")

    # メタ：.race_name（レース名）＋ 帯から 日付/場所/番号
    race_el = soup.select_one(".race_name")
    race_title = race_el.get_text(strip=True).split("|")[0].strip() if race_el else ""

    text_all = soup.get_text(" ", strip=True)
    ymd, place, race_no = extract_basic_meta(text_all)
    if not race_title:
        race_title = f"{place}{race_no}"

    filename = f"{ymd}_{place}_{race_no}.xlsx"
    return rows, filename, race_title


def save_to_desktop(rows, filename, race_title):
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    os.makedirs(desktop, exist_ok=True)
    path = os.path.join(desktop, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "出馬表"

    # 見出し（A1:E1）
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = race_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True, size=18)
    title_cell.fill = PatternFill(start_color="FADADD", end_color="FADADD", fill_type="solid")
    ws.row_dimensions[1].height = 30

    headers = ["馬番","馬名","騎手名","評価","短評"]
    ws.append(headers)

    for umaban, horse, jockey in rows:
        ws.append([umaban, horse, jockey, "", ""])

    # 体裁
    light_blue = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    bold = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    max_row = ws.max_row
    max_col = 5

    for c in range(1, max_col + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = light_blue
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50

    wb.save(path)
    return path


# ====== URL組み立て（任意の簡易ビルダー：失敗時はURL直入れにフォールバック） ======
def build_jra_url(yyyymmdd: str, place: str, race_no: int) -> str | None:
    """
    JRA公式のURLはページ種別や開催回次で揺れるため、確実性は保証しません。
    失敗したら None を返すので、その場合はURL直入れで使ってください。
    """
    # まずはシンプルに検索的なURLパターン（構造変更に弱いので完全フォールバック設計）
    try:
        d = datetime.strptime(yyyymmdd, "%Y%m%d")
    except ValueError:
        return None
    # ここでは汎用性重視で、印刷ページ等に寄せた推測パターンは避けます
    return None  # セレクトからは基本 URL 未使用 → 入力URLを優先推奨


# ====== GUI本体 ======
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("JRA出馬表 取り込みツール（Excel出力）")
        self.geometry("560x260")
        self.resizable(False, False)

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        # 入力群
        row = 0
        ttk.Label(frm, text="取得方法").grid(row=row, column=0, sticky="w")
        self.mode = tk.StringVar(value="select")
        ttk.Radiobutton(frm, text="日付・場所・レースを選択", variable=self.mode, value="select").grid(row=row, column=1, sticky="w")
        ttk.Radiobutton(frm, text="URLを直接入力",     variable=self.mode, value="url").grid(row=row, column=2, sticky="w")

        row += 1
        ttk.Label(frm, text="日付 (YYYYMMDD)").grid(row=row, column=0, sticky="e", pady=6)
        self.ent_date = ttk.Entry(frm, width=16)
        self.ent_date.grid(row=row, column=1, sticky="w", pady=6)
        self.ent_date.insert(0, datetime.now().strftime("%Y%m%d"))

        ttk.Label(frm, text="場所").grid(row=row, column=2, sticky="e")
        self.cmb_place = ttk.Combobox(frm, values=VENUES, width=8, state="readonly")
        self.cmb_place.grid(row=row, column=3, sticky="w")
        self.cmb_place.set("京都")

        row += 1
        ttk.Label(frm, text="レース").grid(row=row, column=0, sticky="e")
        self.cmb_race = ttk.Combobox(frm, values=[f"{i}R" for i in range(1,13)], width=6, state="readonly")
        self.cmb_race.grid(row=row, column=1, sticky="w")
        self.cmb_race.set("11R")

        row += 1
        ttk.Label(frm, text="URL（任意：直入力する場合）").grid(row=row, column=0, sticky="e")
        self.ent_url = ttk.Entry(frm, width=48)
        self.ent_url.grid(row=row, column=1, columnspan=3, sticky="w")

        # 実行ボタン
        row += 1
        self.btn = ttk.Button(frm, text="Excelを作成（デスクトップに保存）", command=self.run_fetch)
        self.btn.grid(row=row, column=0, columnspan=4, sticky="ew", pady=12)

        # ステータス
        row += 1
        self.status = ttk.Label(frm, text="待機中")
        self.status.grid(row=row, column=0, columnspan=4, sticky="w")

        for i in range(4):
            frm.columnconfigure(i, weight=1)

    def run_fetch(self):
        mode = self.mode.get()
        url = self.ent_url.get().strip()

        if mode == "url":
            if not url:
                messagebox.showwarning("入力不足", "URLを入力してください。")
                return
            target_url = url
        else:
            ymd = self.ent_date.get().strip()
            place = self.cmb_place.get().strip()
            race = self.cmb_race.get().strip()
            if not (re.fullmatch(r"\d{8}", ymd) and place in VENUES and re.fullmatch(r"\d{1,2}R", race)):
                messagebox.showwarning("入力値エラー", "日付はYYYYMMDD、場所はリストから、レースは1R〜12Rを選択してください。")
                return
            # URLビルドは確実性がないため、基本はURL直入れを推奨
            built = build_jra_url(ymd, place, int(race[:-1]))
            if built:
                target_url = built
            else:
                # セレクトで URL を組めない場合は案内
                if not url:
                    messagebox.showinfo("URL案内", "JRA公式の出馬表ページURLをコピーして、上のURL欄へ貼り付けてください。")
                    return
                target_url = url

        # 非同期実行
        self.btn.config(state="disabled")
        self.status.config(text="取得中…")
        threading.Thread(target=self._do_fetch, args=(target_url,), daemon=True).start()

    def _do_fetch(self, target_url: str):
        try:
            rows, filename, race_title = fetch_rows_and_meta_by_url(target_url)
            out = save_to_desktop(rows, filename, race_title)
            self._done(f"保存完了：{out}")
        except Exception as e:
            self._done(f"エラー：{e}")

    def _done(self, msg: str):
        # UIスレッドに戻す
        self.after(0, lambda: self._finish_ui(msg))

    def _finish_ui(self, msg: str):
        self.btn.config(state="normal")
        self.status.config(text=msg)
        if msg.startswith("保存完了"):
            messagebox.showinfo("完了", msg)
        else:
            messagebox.showerror("失敗", msg)


if __name__ == "__main__":
    App().mainloop()